
# import os
# from openpyxl import load_workbook, Workbook
# from openai import OpenAI

# # 从环境变量获取API密钥
# API_KEY = "xx"
# # 使用指定的API URL和模型
# client = OpenAI(api_key=API_KEY, base_url="https://api.deepseek.com")

# def split_text(text, max_length=2000):
#     texts = []
#     current = []
#     current_len = 0
#     for line in text.split('\n'):
#         if current_len + len(line) > max_length:
#             texts.append('\n'.join(current))
#             current = [line]
#             current_len = len(line)
#         else:
#             current.append(line)
#             current_len += len(line)
#     if current:
#         texts.append('\n'.join(current))
#     return texts

# def translate_text(text):
#     print(f"正在翻译文本: {text[:50]}...")  # 提示正在翻译的文本片段
#     response = client.chat.completions.create(
#         model="deepseek-chat",
#         messages=[
#             {"role": "user", "content": f"请将以下内容翻译为另一种语言，保持格式不变:\n{text}"},
#         ],
#         temperature=0.7
#     )
#     translated_text = response.choices[0].message.content.strip()
#     print(f"翻译完成: {translated_text[:50]}...")  # 提示翻译完成的文本片段
#     return translated_text

# def process_excel(input_file, output_file):
#     # 检查输入文件
#     if not os.path.isfile(input_file):
#         print(f"错误：文件 {input_file} 不存在")
#         return
    
#     print(f"开始处理文件: {input_file}")  # 提示开始处理文件
    
#     # 创建输出目录
#     output_dir = os.path.dirname(os.path.abspath(output_file))
#     os.makedirs(output_dir, exist_ok=True)
    
#     # 加载源文件和创建目标文件
#     wb_source = load_workbook(input_file)
#     wb_target = Workbook()  # 创建空目标工作簿
    
#     # 处理每个工作表
#     for sheet_name in wb_source.sheetnames:
#         print(f"正在处理工作表: {sheet_name}")  # 提示正在处理的工作表
        
#         ws_source = wb_source[sheet_name]
#         ws_target = wb_target.create_sheet(title=sheet_name)
        
#         # 处理每个单元格
#         for row_idx, row in enumerate(ws_source.iter_rows(values_only=False), start=1):
#             for col_idx, cell in enumerate(row, start=1):
#                 value = cell.value
#                 if isinstance(value, str):
#                     translated_parts = [translate_text(part) for part in split_text(value)]
#                     ws_target.cell(row=row_idx, column=col_idx).value = "\n".join(translated_parts)
#                 else:
#                     ws_target.cell(row=row_idx, column=col_idx).value = value
    
#     # 保存结果
#     wb_target.save(output_file)
#     print(f"翻译完成，结果已保存到 {output_file}")  # 提示翻译完成并保存文件

# if __name__ == "__main__":
#     INPUT_FILE = input("请输入要翻译的Excel文件名: ")
#     OUTPUT_FILE = INPUT_FILE.rsplit('.', 1)[0] + "_fy.xlsx"
#     try:
#         process_excel(INPUT_FILE, OUTPUT_FILE)
#     except Exception as e:
#         print(f"发生错误: {e}")

# 并发处理
import os
from openpyxl import load_workbook, Workbook
from openai import OpenAI
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache

# 从环境变量获取API密钥
API_KEY = "sk-9c4295a93fca4d28953cf4460963975a"

# 使用指定的API URL和模型
client = OpenAI(api_key=API_KEY, base_url="https://api.deepseek.com")

@lru_cache(maxsize=None)
def translate_text(text):
    print(f"正在翻译文本: {text[:50]}...")  # 提示正在翻译的文本片段
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "user", "content": f"请将以下内容翻译为另一种语言，保持格式不变:\n{text}"},
        ],
        temperature=0.7
    )
    translated_text = response.choices[0].message.content.strip()
    print(f"翻译完成: {translated_text[:50]}...")  # 提示翻译完成的文本片段
    return translated_text

def process_cell(value):
    if isinstance(value, str):
        translated_parts = [translate_text(part) for part in split_text(value)]
        return "\n".join(translated_parts)
    else:
        return value

def process_sheet(sheet_name, ws_source, ws_target):
    print(f"正在处理工作表: {sheet_name}")  # 提示正在处理的工作表
    
    for row_idx, row in enumerate(ws_source.iter_rows(values_only=False), start=1):
        for col_idx, cell in enumerate(row, start=1):
            value = cell.value
            ws_target.cell(row=row_idx, column=col_idx).value = process_cell(value)

def split_text(text, max_length=2000):
    texts = []
    current = []
    current_len = 0
    for line in text.split('\n'):
        if current_len + len(line) > max_length:
            texts.append('\n'.join(current))
            current = [line]
            current_len = len(line)
        else:
            current.append(line)
            current_len += len(line)
    if current:
        texts.append('\n'.join(current))
    return texts

def process_excel(input_file, output_file):
    # 检查输入文件
    if not os.path.isfile(input_file):
        print(f"错误：文件 {input_file} 不存在")
        return
    
    print(f"开始处理文件: {input_file}")  # 提示开始处理文件
    
    # 创建输出目录
    output_dir = os.path.dirname(os.path.abspath(output_file))
    os.makedirs(output_dir, exist_ok=True)
    
    # 加载源文件和创建目标文件
    wb_source = load_workbook(input_file)
    wb_target = Workbook()  # 创建空目标工作簿
    
    # 并发处理每个工作表
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = [
            executor.submit(process_sheet, sheet_name, wb_source[sheet_name], wb_target.create_sheet(title=sheet_name))
            for sheet_name in wb_source.sheetnames
        ]
        for future in futures:
            future.result()  # 等待所有任务完成
    
    # 保存结果
    wb_target.save(output_file)
    print(f"翻译完成，结果已保存到 {output_file}")  # 提示翻译完成并保存文件

if __name__ == "__main__":
    INPUT_FILE = input("请输入要翻译的Excel文件名: ")
    OUTPUT_FILE = INPUT_FILE.rsplit('.', 1)[0] + "_fy.xlsx"
    try:
        process_excel(INPUT_FILE, OUTPUT_FILE)
    except Exception as e:
        print(f"发生错误: {e}")